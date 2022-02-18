#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Part of the PsychoPy library
# Copyright (C) 2002-2018 Jonathan Peirce (C) 2019-2022 Open Science Tools Ltd.
# Distributed under the terms of the GNU General Public License (GPL).

import os
import re
import ast
import pickle
import time, datetime
import numpy as np
import pandas as pd

from collections import OrderedDict
from pkg_resources import parse_version

from psychopy import logging, exceptions
from psychopy.tools.filetools import pathToString

try:
    import openpyxl
    if parse_version(openpyxl.__version__) >= parse_version('2.4.0'):
        # openpyxl moved get_column_letter to utils.cell
        from openpyxl.utils.cell import get_column_letter
    else:
        from openpyxl.cell import get_column_letter
    from openpyxl.reader.excel import load_workbook
    haveOpenpyxl = True
except ImportError:
    haveOpenpyxl = False

haveXlrd = False

_nonalphanumeric_re = re.compile(r'\W')  # will match all bad var name chars


def checkValidFilePath(filepath, makeValid=True):
    """Checks whether file path location (e.g. is a valid folder)

    This should also check whether we have write-permissions to the folder
    but doesn't currently do that!

    added in: 1.90.00
    """
    folder = os.path.split(os.path.abspath(filepath))[0]
    if not os.path.isdir(folder):
        os.makedirs(folder)  # spit an error if we fail
    return True


def isValidVariableName(name):
    """Checks whether a certain string could be used as a valid variable.

    Usage::

        OK, msg = isValidVariableName(name)

    >>> isValidVariableName('name')
    (True, '')
    >>> isValidVariableName('0name')
    (False, 'Variables cannot begin with numeric character')
    >>> isValidVariableName('first second')
    (False, 'Variables cannot contain punctuation or spaces')
    >>> isValidVariableName('')
    (False, "Variables cannot be missing, None, or ''")
    >>> isValidVariableName(None)
    (False, "Variables cannot be missing, None, or ''")
    >>> isValidVariableName(23)
    (False, "Variables must be string-like")
    >>> isValidVariableName('a_b_c')
    (True, '')
    """
    if not name:
        return False, "Variables cannot be missing, None, or ''"
    if not isinstance(name, str):
        return False, "Variables must be string-like"
    try:
        name = str(name)  # convert from unicode if possible
    except Exception:
        if type(name) in [str, np.unicode_]:
            msg = ("name %s (type %s) contains non-ASCII characters"
                   " (e.g. accents)")
            raise AttributeError(msg % (name, type(name)))
        else:
            msg = "name %s (type %s) could not be converted to a string"
            raise AttributeError(msg % (name, type(name)))

    if name[0].isdigit():
        return False, "Variables cannot begin with numeric character"
    if _nonalphanumeric_re.search(name):
        return False, "Variables cannot contain punctuation or spaces"
    return True, ''


def _getExcelCellName(col, row):
    """Returns the excel cell name for a row and column (zero-indexed)

    >>> _getExcelCellName(0,0)
    'A1'
    >>> _getExcelCellName(2,1)
    'C2'
    """
    # BEWARE - openpyxl uses indexing at 1, to fit with Excel
    return "%s%i" % (get_column_letter(col + 1), row + 1)


def importTrialTypes(fileName, returnFieldNames=False):
    """importTrialTypes is DEPRECATED (as of v1.70.00)
    Please use `importConditions` for identical functionality.
    """
    logging.warning("importTrialTypes is DEPRECATED (as of v1.70.00). "
                    "Please use `importConditions` for identical "
                    "functionality.")
    return importConditions(fileName, returnFieldNames)


def sliceFromString(sliceString):
    """Convert a text string into a valid slice object
    which can be used as indices for a list or array.

    >>> sliceFromString("0:10")
    slice(0,10,None)
    >>> sliceFromString("0::3")
    slice(0,None,3)
    >>> sliceFromString("-8:")
    slice(-8,None,None)
    """
    sliceArgs = []
    for val in sliceString.split(':'):
        if len(val) == 0:
            sliceArgs.append(None)
        else:
            sliceArgs.append(int(round(float(val))))
            # nb int(round(float(x))) is needed for x='4.3'
    return slice(*sliceArgs)


def indicesFromString(indsString):
    """Convert a text string into a valid list of indices
    """
    # "6"
    try:
        inds = int(round(float(indsString)))
        return [inds]
    except Exception:
        pass
    # "-6::2"
    try:
        inds = sliceFromString(indsString)
        return inds
    except Exception:
        pass
    # "1,4,8"
    try:
        inds = list(eval(indsString))
        return inds
    except Exception:
        pass


def listFromString(val, excludeEmpties=False):
    """Take a string that looks like a list (with commas and/or [] and make
    an actual python list"""
    # was previously called strToList and str2list might have been an option!
    # I'll leave those here for anyone doing a find-in-path for those
    if type(val) == tuple:
        return list(val)
    elif type(val) == list:
        return list(val)  # nothing to do
    elif type(val) != str:
        raise ValueError("listFromString requires a string as its input not {}"
                         .format(repr(val)))
    # try to evaluate with ast (works for "'yes,'no'" or "['yes', 'no']")
    try:
        iterable = ast.literal_eval(val)
        if type(iterable) == tuple:
            iterable = list(iterable)
        return iterable
    except (ValueError, SyntaxError):
        pass  # e.g. "yes, no" won't work. We'll go on and try another way

    val = val.strip()  # in case there are spaces
    if val.startswith(('[', '(')) and val.endswith((']', ')')):
        val = val[1:-1]
    asList = val.split(",")
    if excludeEmpties:
        asList = [this.strip() for this in asList if this]
    else:
        asList = [this.strip() for this in asList]
    return asList


def importConditions(fileName, returnFieldNames=False, selection=""):
    """Imports a list of conditions from an .xlsx, .csv, or .pkl file

    The output is suitable as an input to :class:`TrialHandler`
    `trialList` or to :class:`MultiStairHandler` as a `conditions` list.

    If `fileName` ends with:

    - .csv:  import as a comma-separated-value file
            (header + row x col)
    - .xlsx: import as Excel 2007 (xlsx) files.
            No support for older (.xls) is planned.
    - .pkl:  import from a pickle file as list of lists
            (header + row x col)

    The file should contain one row per type of trial needed and one column
    for each parameter that defines the trial type. The first row should give
    parameter names, which should:

    - be unique
    - begin with a letter (upper or lower case)
    - contain no spaces or other punctuation (underscores are permitted)


    `selection` is used to select a subset of condition indices to be used
    It can be a list/array of indices, a python `slice` object or a string to
    be parsed as either option.
    e.g.:

    - "1,2,4" or [1,2,4] or (1,2,4) are the same
    - "2:5"       # 2, 3, 4 (doesn't include last whole value)
    - "-10:2:"    # tenth from last to the last in steps of 2
    - slice(-10, 2, None)  # the same as above
    - random(5) * 8  # five random vals 0-7

    """

    def _attemptImport(fileName, sep=',', dec='.'):
        """Attempts to import file with specified settings and raises
        ConditionsImportError if fails due to invalid format

        :param filename: str
        :param sep: str indicating the separator for cells (',', ';' etc)
        :param dec: str indicating the decimal point ('.', '.')
        :return: trialList, fieldNames
        """
        if fileName.endswith(('.csv', '.tsv')):
            trialsArr = pd.read_csv(fileName, encoding='utf-8-sig',
                                    sep=sep, decimal=dec)
            for col in trialsArr.columns:
                for row, cell in enumerate(trialsArr[col]):
                    if isinstance(cell, str):
                        tryVal = cell.replace(",", ".")
                        try:
                            trialsArr[col][row] = float(tryVal)
                        except ValueError:
                            pass
            logging.debug(u"Read csv file with pandas: {}".format(fileName))
        elif fileName.endswith(('.xlsx', '.xlsm')):
            trialsArr = pd.read_excel(fileName, engine='openpyxl')
            logging.debug(u"Read Excel file with pandas: {}".format(fileName))
        elif fileName.endswith('.xls'):
            trialsArr = pd.read_excel(fileName, engine='xlrd')
            logging.debug(u"Read Excel file with pandas: {}".format(fileName))
        # then try to convert array to trialList and fieldnames
        unnamed = trialsArr.columns.to_series().str.contains('^Unnamed: ')
        trialsArr = trialsArr.loc[:, ~unnamed]  # clear unnamed cols
        logging.debug(u"Clearing unnamed columns from {}".format(fileName))
        trialList, fieldNames = pandasToDictList(trialsArr)

        return trialList, fieldNames

    def _assertValidVarNames(fieldNames, fileName):
        """screens a list of names as candidate variable names. if all
        names are OK, return silently; else raise  with msg
        """
        fileName = pathToString(fileName)
        if not all(fieldNames):
            msg = ('Conditions file %s: Missing parameter name(s); '
                   'empty cell(s) in the first row?')
            raise exceptions.ConditionsImportError(msg % fileName)
        for name in fieldNames:
            OK, msg = isValidVariableName(name)
            if not OK:
                # tailor message to importConditions
                msg = msg.replace('Variables', 'Parameters (column headers)')
                raise exceptions.ConditionsImportError(
                    'Conditions file %s: %s%s"%s"' %
                    (fileName, msg, os.linesep * 2, name))

    if fileName in ['None', 'none', None]:
        if returnFieldNames:
            return [], []
        return []
    if not os.path.isfile(fileName):
        msg = 'Conditions file not found: %s'
        raise ValueError(msg % os.path.abspath(fileName))

    def pandasToDictList(dataframe):
        """Convert a pandas dataframe to a list of dicts.
        This helper function is used by csv or excel imports via pandas
        """
        # convert the resulting dataframe to a numpy recarray
        trialsArr = dataframe.to_records(index=False)
        # Check for new line characters in strings, and replace escaped characters
        for record in trialsArr:
            for idx, element in enumerate(record):
                if isinstance(element, str):
                    record[idx] = element.replace('\\n', '\n')
        if trialsArr.shape == ():
            # convert 0-D to 1-D with one element:
            trialsArr = trialsArr[np.newaxis]
        fieldNames = list(trialsArr.dtype.names)
        _assertValidVarNames(fieldNames, fileName)

        # convert the record array into a list of dicts
        trialList = []
        for trialN, trialType in enumerate(trialsArr):
            thisTrial = OrderedDict()
            for fieldN, fieldName in enumerate(fieldNames):
                val = trialsArr[trialN][fieldN]

                if isinstance(val, str):
                    if val.startswith('[') and val.endswith(']'):
                        # val = eval('%s' %unicode(val.decode('utf8')))
                        val = eval(val)
                elif type(val) == np.string_:
                    val = str(val.decode('utf-8-sig'))
                    # if it looks like a list, convert it:
                    if val.startswith('[') and val.endswith(']'):
                        # val = eval('%s' %unicode(val.decode('utf8')))
                        val = eval(val)
                elif np.isnan(val):
                    val = None
                thisTrial[fieldName] = val
            trialList.append(thisTrial)
        return trialList, fieldNames

    if (fileName.endswith(('.csv', '.tsv'))
            or (fileName.endswith(('.xlsx', '.xls', '.xlsm')) and haveXlrd)):
        if fileName.endswith(('.csv', '.tsv', '.dlm')):  # delimited text file
            for sep, dec in [ (',', '.'), (';', ','),  # most common in US, EU
                              ('\t', '.'), ('\t', ','), (';', '.')]:
                try:
                    trialList, fieldNames = _attemptImport(fileName=fileName,
                                                           sep=sep, dec=dec)
                    break  # seems to have worked
                except exceptions.ConditionsImportError as e:
                    continue  # try a different format
        else:
            trialList, fieldNames = _attemptImport(fileName=fileName)

    elif fileName.endswith(('.xlsx','.xlsm')):  # no xlsread so use openpyxl
        if not haveOpenpyxl:
            raise ImportError('openpyxl or xlrd is required for loading excel '
                              'files, but neither was found.')

        # data_only was added in 1.8
        if parse_version(openpyxl.__version__) < parse_version('1.8'):
            wb = load_workbook(filename=fileName)
        else:
            wb = load_workbook(filename=fileName, data_only=True)
        ws = wb.worksheets[0]

        logging.debug(u"Read excel file with openpyxl: {}".format(fileName))
        try:
            # in new openpyxl (2.3.4+) get_highest_xx is deprecated
            nCols = ws.max_column
            nRows = ws.max_row
        except Exception:
            # version openpyxl 1.5.8 (in Standalone 1.80) needs this
            nCols = ws.get_highest_column()
            nRows = ws.get_highest_row()

        # get parameter names from the first row header
        fieldNames = []
        rangeCols = list(range(nCols))
        for colN in range(nCols):
            if parse_version(openpyxl.__version__) < parse_version('2.0'):
                fieldName = ws.cell(_getExcelCellName(col=colN, row=0)).value
            else:
                # From 2.0, cells are referenced with 1-indexing: A1 == cell(row=1, column=1)
                fieldName = ws.cell(row=1, column=colN + 1).value
            if fieldName:
                # If column is named, add its name to fieldNames
                fieldNames.append(fieldName)
            else:
                # Otherwise, ignore the column
                rangeCols.remove(colN)
        _assertValidVarNames(fieldNames, fileName)

        # loop trialTypes
        trialList = []
        for rowN in range(1, nRows):  # skip header first row
            thisTrial = {}
            for colN in rangeCols:
                if parse_version(openpyxl.__version__) < parse_version('2.0'):
                    val = ws.cell(_getExcelCellName(col=colN, row=0)).value
                else:
                    # From 2.0, cells are referenced with 1-indexing: A1 == cell(row=1, column=1)
                    val = ws.cell(row=rowN + 1, column=colN + 1).value
                # if it looks like a list or tuple, convert it
                if (isinstance(val, str) and
                        (val.startswith('[') and val.endswith(']') or
                                 val.startswith('(') and val.endswith(')'))):
                    val = eval(val)
                # if it has any line breaks correct them
                if isinstance(val, str):
                    val = val.replace('\\n', '\n')
                # Convert from eu style decimals: replace , with . and try to make it a float
                if isinstance(val, str):
                    tryVal = val.replace(",", ".")
                    try:
                        val = float(tryVal)
                    except ValueError:
                        pass
                fieldName = fieldNames[colN]
                thisTrial[fieldName] = val
            trialList.append(thisTrial)

    elif fileName.endswith('.pkl'):
        f = open(fileName, 'rb')
        # Converting newline characters.
        # 'b' is necessary in Python3 because byte object is
        # returned when file is opened in binary mode.
        buffer = f.read().replace(b'\r\n',b'\n').replace(b'\r',b'\n')
        try:
            trialsArr = pickle.loads(buffer)
        except Exception:
            raise IOError('Could not open %s as conditions' % fileName)
        f.close()
        trialList = []
        # In Python3, strings returned by pickle() are unhashable so we have to
        # convert them to str.
        trialsArr = [[str(item) if isinstance(item, str) else item
                      for item in row] for row in trialsArr]
        fieldNames = trialsArr[0]  # header line first
        _assertValidVarNames(fieldNames, fileName)
        for row in trialsArr[1:]:
            thisTrial = {}
            for fieldN, fieldName in enumerate(fieldNames):
                # type is correct, being .pkl
                thisTrial[fieldName] = row[fieldN]
            trialList.append(thisTrial)
    else:
        raise IOError('Your conditions file should be an '
                      'xlsx, csv, dlm, tsv or pkl file')

    # if we have a selection then try to parse it
    if isinstance(selection, str) and len(selection) > 0:
        selection = indicesFromString(selection)
        if not isinstance(selection, slice):
            for n in selection:
                try:
                    assert n == int(n)
                except AssertionError:
                    raise TypeError("importConditions() was given some "
                                    "`indices` but could not parse them")

    # the selection might now be a slice or a series of indices
    if isinstance(selection, slice):
        trialList = trialList[selection]
    elif len(selection) > 0:
        allConds = trialList
        trialList = []
        print(selection)
        print(len(allConds))
        for ii in selection:
            trialList.append(allConds[int(ii)])

    logging.exp('Imported %s as conditions, %d conditions, %d params' %
                (fileName, len(trialList), len(fieldNames)))
    if returnFieldNames:
        return (trialList, fieldNames)
    else:
        return trialList


def createFactorialTrialList(factors):
    """Create a trialList by entering a list of factors with names (keys)
    and levels (values) it will return a trialList in which all factors
    have been factorially combined (so for example if there are two factors
    with 3 and 5 levels the trialList will be a list of 3*5 = 15, each
    specifying the values for a given trial

    Usage::

        trialList = createFactorialTrialList(factors)

    :Parameters:

        factors : a dictionary with names (keys) and levels (values) of the
            factors

    Example::

        factors={"text": ["red", "green", "blue"],
                 "letterColor": ["red", "green"],
                 "size": [0, 1]}
        mytrials = createFactorialTrialList(factors)
    """

    # the first step is to place all the factorial combinations in a list of
    # lists
    tempListOfLists = [[]]
    for key in factors:
        # this takes the levels of each factor as a set of values
        # (a list) at a time
        alist = factors[key]
        tempList = []
        for value in alist:
            # now we loop over the values in a given list,
            # and add each value of the other lists
            for iterList in tempListOfLists:
                tempList.append(iterList + [key, value])
        tempListOfLists = tempList

    # this second step is so we can return a list in the format of trialList
    trialList = []
    for atrial in tempListOfLists:
        keys = atrial[0::2]  # the even elements are keys
        values = atrial[1::2]  # the odd elements are values
        atrialDict = {}
        for i in range(len(keys)):
            # this combines the key with the value
            atrialDict[keys[i]] = values[i]
        # append one trial at a time to the final trialList
        trialList.append(atrialDict)

    return trialList


def bootStraps(dat, n=1):
    """Create a list of n bootstrapped resamples of the data

    SLOW IMPLEMENTATION (Python for-loop)

    Usage:
        ``out = bootStraps(dat, n=1)``

    Where:
        dat
            an NxM or 1xN array (each row is a different condition, each
            column is a different trial)
        n
            number of bootstrapped resamples to create

        out
            - dim[0]=conditions
            - dim[1]=trials
            - dim[2]=resamples
    """
    dat = np.asarray(dat)
    if len(dat.shape) == 1:
        # have presumably been given a series of data for one stimulus
        # adds a dimension (arraynow has shape (1,Ntrials))
        dat = np.array([dat])

    nTrials = dat.shape[1]
    # initialise a matrix to store output
    resamples = np.zeros(dat.shape + (n,), dat.dtype)
    rand = np.random.rand
    for stimulusN in range(dat.shape[0]):
        thisStim = dat[stimulusN, :]  # fetch data for this stimulus
        for sampleN in range(n):
            indices = np.floor(nTrials * rand(nTrials)).astype('i')
            resamples[stimulusN, :, sampleN] = np.take(thisStim, indices)
    return resamples


def functionFromStaircase(intensities, responses, bins=10):
    """Create a psychometric function by binning data from a staircase
    procedure. Although the default is 10 bins Jon now always uses 'unique'
    bins (fewer bins looks pretty but leads to errors in slope estimation)

    usage::

        intensity, meanCorrect, n = functionFromStaircase(intensities,
                                                          responses, bins)

    where:
            intensities
                are a list (or array) of intensities to be binned

            responses
                are a list of 0,1 each corresponding to the equivalent
                intensity value

            bins
                can be an integer (giving that number of bins) or 'unique'
                (each bin is made from aa data for exactly one intensity
                value)

            intensity
                a numpy array of intensity values (where each is the center
                of an intensity bin)

            meanCorrect
                a numpy array of mean % correct in each bin

            n
                a numpy array of number of responses contributing to each mean
    """
    # convert to arrays
    try:
        # concatenate if multidimensional
        intensities = np.concatenate(intensities)
        responses = np.concatenate(responses)
    except Exception:
        intensities = np.array(intensities)
        responses = np.array(responses)

    # sort the responses
    sort_ii = np.argsort(intensities)
    sortedInten = np.take(intensities, sort_ii)
    sortedResp = np.take(responses, sort_ii)

    binnedResp = []
    binnedInten = []
    nPoints = []
    if bins == 'unique':
        intensities = np.round(intensities, decimals=8)
        uniqueIntens = np.unique(intensities)
        for thisInten in uniqueIntens:
            theseResps = responses[intensities == thisInten]
            binnedInten.append(thisInten)
            binnedResp.append(np.mean(theseResps))
            nPoints.append(len(theseResps))
    else:
        pointsPerBin = len(intensities)/bins
        for binN in range(bins):
            start = int(round(binN * pointsPerBin))
            stop = int(round((binN + 1) * pointsPerBin))
            thisResp = sortedResp[start:stop]
            thisInten = sortedInten[start:stop]
            binnedResp.append(np.mean(thisResp))
            binnedInten.append(np.mean(thisInten))
            nPoints.append(len(thisInten))

    return binnedInten, binnedResp, nPoints


def getDateStr(format="%Y-%m-%d_%Hh%M.%S.%f", fractionalSecondDigits=3):
    """Uses ``datetime.now().strftime(format)``_ to generate a string
    based on ISO 8601 but made safe for filename use::

        "2022-01-14_18h35.05.386"

    represents 14th Jan 2022 at 6:35pm with 5 sec and 386 ms

    This is often useful appended to data filenames to provide unique names.

    Parameters
    ----------
    format : str
        default="%Y-%m-%d_%Hh%M.%S.%f"
    fractionalSecondDigits : int
        An integer value 1-6 indicating the number of digits of fractional
        seconds to include if the `%f` parameter is included in the format.
        This would normally give 6 digits (microseconds) but to get just
        milliseconds you can set fractionalSecondDigits=3

    """
    now = datetime.datetime.now()
    microsecs = now.strftime("%f")
    nowStr = now.strftime(format)
    if "%f" in format and (
            fractionalSecondDigits < 1
            or int(fractionalSecondDigits) != fractionalSecondDigits
    ):
        raise TypeError("fractionalSecondDigits argument to getDateStr should "
                        f"be an integer greater than 1, not {fractionalSecondDigits}")
    elif  "%f" in format and fractionalSecondDigits > len(microsecs):
        logging.warning("fractionalSecondDigits argument to getDateStr requested "
                        f"{fractionalSecondDigits} digits but only {len(microsecs)} "
                        f"are available. Truncating to {len(microsecs)}.")
    elif "%f" in format:
        nowStr = nowStr.replace(
            microsecs, microsecs[:int(fractionalSecondDigits)],
        )
    return nowStr
